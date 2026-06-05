import yaml
from .md_file_handler import __cell, _parse_cell, _parse_dict
try:
    import openpyxl
except ImportError:
    openpyxl = None

EMPTY = object()

def read_excel(filename, sheet=None, **kwargs):

    assert openpyxl is not None, ImportError("openpyxl required for xlsx based metadata")

    wb = openpyxl.load_workbook(
                filename,
                data_only=True,
                read_only=True,
                **kwargs,
            )

    ws = wb.active if sheet is None else wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    source_rows = {}

    # parse each cell individually.
    # this will honor yml syntax like comments and converts to python basic types.
    cleaned = []
    for source_row, row in enumerate(rows, start=1):
        row = [_parse_cell(v) for v in row]
        # remove empty cells at the end of rows.
        while row and row[-1] is EMPTY:
            row.pop()

        if row:
            source_rows[len(cleaned)] = source_row
            cleaned.append(row)

    rows = cleaned
    return _parse_dict(source_rows, rows, row=0)[0]


def to_excel(metadata, filename, title="metadata", **kwargs):

        assert openpyxl is not None, ImportError("openpyxl required for xlsx based metadata")

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Side, Border

        wb, ws = Workbook(), Workbook().active
        ws = wb.active
        ws.title = title

        fills = ["FFFFFF", "F2F2F2", "EAF1FF", "EAF7EA"]
        colors = ["000000", "2F5597", "38761D", "7F3F98"]

        def style(r, d):
            ws.cell(r, d + 1).fill = PatternFill("solid", fgColor=fills[d % 4])
            ws.cell(r, d + 1).font = Font(bold=True)
            ws.cell(r, d + 1).border = Border(left=Side("thick", colors[d % 4]))

        def _cell(r, d, k):
            ws.cell(r, d).number_format = "@" #force to string??
            try:
                ws.cell(r, d, __cell(k))
            except:
                ws.cell(r, d, str(__cell(k)))

        def write(d:dict, r:int, depth:int, worksheet):
            for k, v in d.items():

                style(r, depth)                     # key row
                _cell(r, depth+1, k)                # value
                worksheet.row_dimensions[r].outlineLevel = depth

                if isinstance(v, dict): #write dictionary block starting on next row
                    r = write(v, r+1, depth + 1, worksheet=worksheet)

                elif isinstance(v, list) and (not v or not isinstance(v[0], list)):
                    _cell(r, depth+2, ':1D')
                    r += 1
                    for x in v:
                        _cell(r, depth+2, x)
                        worksheet.row_dimensions[r].outlineLevel = depth + 1
                        r += 1

                elif isinstance(v, list):
                    _cell(r, depth+2, ':2D')
                    r += 1
                    for row in v:
                        for i, x in enumerate(row):
                            _cell(r, depth+2+i, x)
                        worksheet.row_dimensions[r].outlineLevel = depth + 1
                        r += 1

                else: #scalar
                    _cell(r, depth+2, v)
                    r += 1

            return r

        write(metadata, 1, 0, worksheet=ws)
        ws.sheet_properties.outlinePr.summaryBelow = True
        wb.save(filename)